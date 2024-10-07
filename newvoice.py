from tkinter import *
#import voice_support
import os.path
import pyttsx3
import speech_recognition as sr
import setuptools.dist
from pydub.playback import play
from pydub import AudioSegment
import qrcode
import time
from tkinter import messagebox
import datetime
import openpyxl
from openpyxl import Workbook 
def convert_spech1():
    r=sr.Recognizer()
    with sr.Microphone() as src:
    
         print ('say smoething ....')
         r.adjust_for_ambient_noise(src,duration=1)
         myaudio=r.listen(src)
         mytext=r.recognize_google(myaudio,language='ar')
         mytext=mytext.lower()
       
    Entry1.insert(0,mytext)
def convert_spech2():
    r=sr.Recognizer()
    with sr.Microphone() as src:
    
         print ('say smoething ....')
         r.adjust_for_ambient_noise(src,duration=1)
         myaudio=r.listen(src)
         mytext=r.recognize_google(myaudio,language='ar')
         mytext=mytext.lower()
       
    Entry2.insert(0,mytext)    
def convert_spech3():
    r=sr.Recognizer()
    with sr.Microphone() as src:
    
         print ('say smoething ....')
         r.adjust_for_ambient_noise(src,duration=1)
         myaudio=r.listen(src)
         mytext=r.recognize_google(myaudio,language='ar')
         mytext=mytext.lower()
       
    Entry3.insert(0,mytext)    
def welcom():
    music=AudioSegment.from_mp3("1.mp3")
    play(music)
def insertxlsx():
        
        excel=openpyxl.load_workbook('emp.xlsx')
        file=excel.active
        
        file.cell(column=2,row=1,value=Entry1.get())
        file.cell(column=3,row=1,value=Entry2.get())
        
        
        excel.save('emp.xlsx')

win=Tk()
win.geometry("518x450+383+106")
win.iconbitmap()
win.title("الادخال الصوتي")
l_frame=Frame(win,bg='silver',width=600,height=550)
l_frame.place(x=1,y=1)
        
titl_frame=Label(l_frame,text='  ادخل الحقول من خلال الصوت',font=('Tajawal 13'),fg='white',bg='#5F7161',width=70)
titl_frame.place(x=0,y=0)

      
        
imgbuton1=PhotoImage(file="C:/Users/rami/vscode_python_test/market/img/2.png")
imgbuton2=PhotoImage(file="C:/Users/rami/vscode_python_test/market/img/3.png")
Entry1=Entry(win,font=('tajawal,12'),justify=CENTER)
Entry1.place(relx=0.150, rely=0.6, height=30, relwidth=0.375)
buton1=Button(win,width=250,command=convert_spech1,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=200,text='اضغط هنا وسجل الصوت',compound=TOP)
buton1.place(relx=0.550, rely=0.6, height=30, relwidth=0.25)  
Label1=Label(win,text="  اسم الموظف",bg='silver',fg='white')
Label1.place(relx=0.035, rely=0.622, height=21, width=60)


buton2=Button(win,width=250,bg='#EFEAD8',command=convert_spech2,bd=1,relief=SOLID,cursor='hand2',height=200,text='اضغط هنا وسجل الصوت',compound=TOP)
buton2.place(relx=0.550, rely=0.711, height=30, relwidth=0.25) 
Entry2=Entry(win,font=('tajawal,12'),justify=CENTER)
Entry2.place(relx=0.150, rely=0.711, height=30, relwidth=0.375)
Label2=Label(win,text="   العنوان",bg='silver',fg='white')
Label2.place(relx=0.035, rely=0.733, height=21, width=60)

buton3=Button(win,width=250,bg='#EFEAD8',command=convert_spech3,bd=1,relief=SOLID,cursor='hand2',height=200,text='اضغط هنا وسجل الصوت',compound=TOP)
buton3.place(relx=0.550, rely=0.8, height=30, relwidth=0.25)
Entry3=Entry(win,font=('tajawal,12'),justify=CENTER)
Entry3.place(relx=0.150, rely=0.8, height=30, relwidth=0.375)
Label3=Label(win,text="   تاريخ الولادة",bg='silver',fg='white')
Label3.place(relx=0.035, rely=0.822, height=21, width=60)
imglable4=PhotoImage(file="c:/Users/rami/vscode_python_test/tkdesigner_exam/image/1.png")
Label4= Label(win,image=imglable4)
Label4.place(relx=0.058, rely=0.044, height=161, width=454)

imgbutton4 = PhotoImage(file="C:/Users/rami/vscode_python_test/tkdesigner_exam/image/3.png")
Button4= Button(win,text='😄',image=imgbutton4,command=insertxlsx)
Button4.place(relx=0.8, rely=0.6, height=100, relwidth=0.2)       
welcom()
win.mainloop()
